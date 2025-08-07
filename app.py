import streamlit as st
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
import pandas as pd
import os
import tempfile
import zipfile
from io import BytesIO
import platform

def carregar_fontes():
    """Carrega as fontes com fallback robusto para diferentes sistemas"""
    
    # Dicionário de fontes para retornar
    fontes = {}
    
    # Lista de possíveis caminhos para fontes Arial
    caminhos_arial = []
    caminhos_arial_bold = []
    
    if platform.system() == "Windows":
        caminhos_arial = [
            "C:/Windows/Fonts/arial.ttf",
            "C:/Windows/Fonts/Arial.ttf",
            "arial.ttf"
        ]
        caminhos_arial_bold = [
            "C:/Windows/Fonts/arialbd.ttf",
            "C:/Windows/Fonts/ARIALBD.TTF",
            "arialbd.ttf"
        ]
    elif platform.system() == "Darwin":  # macOS
        caminhos_arial = [
            "/System/Library/Fonts/Arial.ttf",
            "/Library/Fonts/Arial.ttf",
            "arial.ttf"
        ]
        caminhos_arial_bold = [
            "/System/Library/Fonts/Arial Bold.ttf",
            "/Library/Fonts/Arial Bold.ttf",
            "arialbd.ttf"
        ]
    else:  # Linux e outros
        caminhos_arial = [
            "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/System/Library/Fonts/Arial.ttf",
            "arial.ttf"
        ]
        caminhos_arial_bold = [
            "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
            "/System/Library/Fonts/Arial Bold.ttf",
            "arialbd.ttf"
        ]
    
    def encontrar_fonte(caminhos, tamanho):
        """Tenta encontrar uma fonte válida nos caminhos fornecidos"""
        for caminho in caminhos:
            try:
                return ImageFont.truetype(caminho, tamanho)
            except (OSError, IOError):
                continue
        # Se nenhuma fonte for encontrada, usa a padrão
        return ImageFont.load_default()
    
    # Carregar todas as fontes necessárias
    try:
        fontes['fonte_pequena'] = encontrar_fonte(caminhos_arial, 14)
        fontes['fonte_media'] = encontrar_fonte(caminhos_arial, 35)
        fontes['fonte_media_b'] = encontrar_fonte(caminhos_arial_bold, 28)
        fontes['fonte_vista'] = encontrar_fonte(caminhos_arial_bold, 26)
        fontes['fonte_parcela'] = encontrar_fonte(caminhos_arial_bold, 38)
        fontes['fonte_p'] = encontrar_fonte(caminhos_arial_bold, 20)
        fontes['fonte_a'] = encontrar_fonte(caminhos_arial_bold, 40)
        fontes['fonte_valor'] = encontrar_fonte(caminhos_arial_bold, 60)
        fontes['fonte_valor_de'] = encontrar_fonte(caminhos_arial_bold, 45)
        
        return fontes
    except Exception as e:
        st.error(f"Erro ao carregar fontes: {e}")
        # Fallback total - usar fonte padrão para tudo
        fonte_default = ImageFont.load_default()
        return {
            'fonte_pequena': fonte_default,
            'fonte_media': fonte_default,
            'fonte_media_b': fonte_default,
            'fonte_vista': fonte_default,
            'fonte_parcela': fonte_default,
            'fonte_p': fonte_default,
            'fonte_a': fonte_default,
            'fonte_valor': fonte_default,
            'fonte_valor_de': fonte_default
        }

def gerar_cartazes(planilha_path, imagem_base_path, pasta_saida):
    """Função que encapsula o código original de geração de cartazes"""
    
    os.makedirs(pasta_saida, exist_ok=True)
    wb = load_workbook(planilha_path)
    ws = wb.active

    # Carregar fontes com sistema robusto
    fontes = carregar_fontes()
    
    cartazes_gerados = []
    
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        codigo, descricao, preco_de, preco_por, parcela, filial, Defeito, Tratativa, Armazem = row

        img = Image.open(imagem_base_path).convert("RGB")
        draw = ImageDraw.Draw(img)

        # Código original mantido com as fontes carregadas
        draw.text((85, 350), "DE :", font=fontes['fonte_media_b'], fill="black")
        
        preco_de_txt = f"R$ {preco_de:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        draw.text((148, 336), preco_de_txt, font=fontes['fonte_valor_de'], fill="black")
        draw.line([(150, 390), (400, 320)], fill="red", width=5)
        draw.line([(146, 326), (371, 400)], fill="red", width=5)

        draw.text((47, 430), "POR :", font=fontes['fonte_media_b'], fill="black")
        
        preco_por_txt = f"R$ {preco_por:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        draw.text((129, 405), preco_por_txt, font=fontes['fonte_valor'], fill="red")

        draw.text((425, 480), "À VISTA", font=fontes['fonte_vista'], fill="black")
        draw.text((44 , 542), "OU 10X\nNO CARTÃO :", font=fontes['fonte_p'], fill="black")
        
        parcela_txt = f"R$ {parcela:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        draw.text((180 , 550), parcela_txt, font=fontes['fonte_parcela'], fill="red")

        draw.text((24, 679), f"FILIAL-{filial}", font=fontes['fonte_pequena'], fill="black")
        draw.text((24, 706), str(codigo), font=fontes['fonte_pequena'], fill="black")
        draw.text((140, 706), str(descricao)[:55], font=fontes['fonte_pequena'], fill="black")
        draw.text((27, 730), str(Defeito), font=fontes['fonte_pequena'], fill="black")
        draw.text((134, 250), str(Tratativa), font=fontes['fonte_a'], fill="black")
        draw.text((380, 679), str(Armazem), font=fontes['fonte_pequena'], fill="black")

        nome_arquivo = f"{pasta_saida}/cartaz_{str(codigo)}.png"
        img.save(nome_arquivo)
        cartazes_gerados.append(nome_arquivo)
        
    return cartazes_gerados

def gerar_pdf(pasta_saida):
    """Função que encapsula o código original de geração do PDF"""
    
    arquivos_png = sorted([f for f in os.listdir(pasta_saida) if f.endswith('.png')])
    imagens = []

    for arquivo in arquivos_png:
        caminho_imagem = os.path.join(pasta_saida, arquivo)
        img = Image.open(caminho_imagem).convert('RGB')
        imagens.append(img)

    if imagens:
        caminho_pdf = os.path.join(pasta_saida, 'cartazes_unificados.pdf')
        imagens[0].save(caminho_pdf, save_all=True, append_images=imagens[1:])
        return caminho_pdf
    else:
        return None

def verificar_sistema():
    """Mostra informações do sistema para debug"""
    info = {
        "Sistema Operacional": platform.system(),
        "Versão": platform.release(),
        "Arquitetura": platform.architecture()[0]
    }
    
    # Verificar se conseguimos acessar algumas fontes comuns
    fontes_encontradas = []
    caminhos_teste = []
    
    if platform.system() == "Windows":
        caminhos_teste = ["C:/Windows/Fonts/arial.ttf", "C:/Windows/Fonts/arialbd.ttf"]
    elif platform.system() == "Darwin":
        caminhos_teste = ["/System/Library/Fonts/Arial.ttf"]
    else:
        caminhos_teste = ["/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"]
    
    for caminho in caminhos_teste:
        if os.path.exists(caminho):
            fontes_encontradas.append(caminho)
    
    return info, fontes_encontradas

def main():
    st.set_page_config(
        page_title="Gerador de Cartazes",
        page_icon="🏷️",
        layout="centered"
    )
    
    st.title("🏷️ Gerador de Cartazes de Preço")
    st.markdown("---")
    
    # Debug do sistema (opcional)
    with st.expander("🔧 Informações do Sistema"):
        info_sistema, fontes_encontradas = verificar_sistema()
        st.json(info_sistema)
        if fontes_encontradas:
            st.success(f"✅ Fontes encontradas: {len(fontes_encontradas)}")
            for fonte in fontes_encontradas:
                st.text(f"• {fonte}")
        else:
            st.warning("⚠️ Nenhuma fonte específica encontrada. Usando fontes padrão.")
    
    # Upload da planilha
    st.subheader("📊 Upload da Planilha")
    uploaded_file = st.file_uploader(
        "Selecione o arquivo Excel (.xlsx)",
        type=['xlsx'],
        help="Faça upload da planilha com os dados dos produtos"
    )
    
    # Upload da imagem base
    st.subheader("🖼️ Upload da Imagem Base")
    uploaded_image = st.file_uploader(
        "Selecione a imagem base do cartaz (.png)",
        type=['png'],
        help="Faça upload da imagem modelo que será usada como base"
    )
    
    if uploaded_file is not None and uploaded_image is not None:
        
        # Criar diretórios temporários
        with tempfile.TemporaryDirectory() as temp_dir:
            
            # Salvar arquivos temporariamente
            planilha_path = os.path.join(temp_dir, "planilha.xlsx")
            imagem_path = os.path.join(temp_dir, "modelo.png")
            pasta_saida = os.path.join(temp_dir, "cartazes_prontos")
            
            with open(planilha_path, "wb") as f:
                f.write(uploaded_file.getbuffer())
            
            with open(imagem_path, "wb") as f:
                f.write(uploaded_image.getbuffer())
            
            # Mostrar preview da planilha
            try:
                wb = load_workbook(planilha_path)
                ws = wb.active
                
                st.subheader("📋 Preview dos Dados")
                
                # Definir os cabeçalhos esperados
                headers = ["código", "descrição", "preço_de", "preço_por", "parcela", "filial", "Defeito", "Tratativa", "Armazém"]
                
                # Criar uma lista com os dados da planilha
                dados = []
                for row in ws.iter_rows(min_row=2, max_row=6, values_only=True):  # Mostrar apenas 5 linhas
                    dados.append(row)
                
                # Mostrar tabela
                if dados:
                    import pandas as pd
                    df_preview = pd.DataFrame(dados, columns=headers)
                    st.dataframe(df_preview)
                    total_rows = ws.max_row - 1  # -1 porque não contamos o cabeçalho
                    st.info(f"Total de produtos na planilha: {total_rows}")
                    
                    # Botão para gerar cartazes
                    if st.button("🚀 Gerar Cartazes", type="primary"):
                        
                        with st.spinner("Gerando cartazes..."):
                            try:
                                # Gerar cartazes
                                cartazes_gerados = gerar_cartazes(planilha_path, imagem_path, pasta_saida)
                                
                                if cartazes_gerados:
                                    st.success(f"✅ {len(cartazes_gerados)} cartazes gerados com sucesso!")
                                    
                                    # Gerar PDF
                                    with st.spinner("Criando PDF..."):
                                        caminho_pdf = gerar_pdf(pasta_saida)
                                    
                                    if caminho_pdf:
                                        st.success("📄 PDF gerado com sucesso!")
                                        
                                        # Mostrar preview do PDF
                                        st.subheader("👀 Visualizar PDF")
                                        
                                        with open(caminho_pdf, "rb") as pdf_file:
                                            pdf_bytes = pdf_file.read()
                                            
                                            # Botão de download
                                            st.download_button(
                                                label="📥 Baixar PDF dos Cartazes",
                                                data=pdf_bytes,
                                                file_name="cartazes_unificados.pdf",
                                                mime="application/pdf",
                                                type="primary"
                                            )
                                            
                                            # Mostrar PDF incorporado
                                            st.markdown("### Preview do PDF:")
                                            st.write("Clique no botão acima para baixar o arquivo PDF completo.")
                                        
                                        # Opção de baixar cartazes individuais em ZIP
                                        st.subheader("📦 Download Individual")
                                        
                                        # Criar ZIP com todos os PNGs
                                        zip_buffer = BytesIO()
                                        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                                            for cartaz in cartazes_gerados:
                                                zip_file.write(cartaz, os.path.basename(cartaz))
                                        
                                        zip_buffer.seek(0)
                                        
                                        st.download_button(
                                            label="📥 Baixar Cartazes Individuais (ZIP)",
                                            data=zip_buffer.getvalue(),
                                            file_name="cartazes_individuais.zip",
                                            mime="application/zip"
                                        )
                                    
                                    else:
                                        st.error("❌ Erro ao gerar PDF")
                                else:
                                    st.error("❌ Nenhum cartaz foi gerado")
                                    
                            except Exception as e:
                                st.error(f"❌ Erro ao processar arquivos: {str(e)}")
                                st.exception(e)  # Mostra stack trace completo para debug
                
                else:
                    st.warning("⚠️ Planilha vazia ou sem dados válidos")
                    
            except Exception as e:
                st.error(f"❌ Erro ao ler planilha: {str(e)}")
    
    else:
        st.info("👆 Faça upload da planilha Excel e da imagem base para começar")
        
        # Instruções
        with st.expander("📖 Instruções de Uso"):
            st.markdown("""
            **Como usar este app:**
            
            1. **Upload da Planilha**: Faça upload do arquivo Excel (.xlsx) contendo os dados dos produtos
            2. **Upload da Imagem**: Faça upload da imagem modelo (.png) que será usada como base dos cartazes
            3. **Visualizar**: Confira os dados da planilha na tabela de preview
            4. **Gerar**: Clique em "Gerar Cartazes" para processar os dados
            5. **Baixar**: Faça download do PDF unificado ou dos cartazes individuais em ZIP
            
            **Formato esperado da planilha:**
            - Coluna A: código
            - Coluna B: descrição  
            - Coluna C: preço_de
            - Coluna D: preço_por
            - Coluna E: parcela
            - Coluna F: filial
            - Coluna G: Defeito
            - Coluna H: Tratativa
            - Coluna I: Armazém
            
            **⚠️ Problemas com fontes?**
            - O app tentará usar Arial automaticamente
            - Se não encontrar, usará fontes padrão do sistema
            - Verifique as "Informações do Sistema" acima para diagnóstico
            """)

if __name__ == "__main__":
    main()

